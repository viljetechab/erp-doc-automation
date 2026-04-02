"""Service for importing articles from CSV/XLSX into the articles table.

Upserts by ``article_number`` — new articles are inserted, existing ones
are updated with non-null values from the import file. This ensures
the database stays current without losing data from prior imports.
"""

from __future__ import annotations

import csv
import io
import re
from decimal import Decimal, InvalidOperation
from typing import Any

import structlog
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import AppError
from app.models.article import Article
from app.schemas.article_import import ArticleImportResponse

logger = structlog.get_logger(__name__)

_SWEDISH_TO_DB: dict[str, str] = {
    "article_number": "article_number",
    "article_name": "artikelbenamning",
    "artikelbenamning": "artikelbenamning",
    "artikeltyp": "artikel_typ_id",
    "artikel_typ_id": "artikel_typ_id",
    "standardpris": "standardpris",
    "totalt saldo": "saldo_varde",
    "saldo_varde": "saldo_varde",
    "saldohanteras": "saldohanteras",
    "artikelkategori": "artikel_kategori_id",
    "artikel_kategori_id": "artikel_kategori_id",
    "artikelkod": "artikel_kod_id",
    "artikel_kod_id": "artikel_kod_id",
    "varugrupp": "varugrupp_id",
    "varugrupp_id": "varugrupp_id",
    "ursprungsland": "ursprungsland",
    "status": "artikel_status_id",
    "artikel_status_id": "artikel_status_id",
    "nettovikt": "nettovikt_varde",
    "nettovikt_varde": "nettovikt_varde",
    "fast vikt": "fast_vikt",
    "fast_vikt": "fast_vikt",
    "enhet": "enhet_id",
    "enhet_id": "enhet_id",
    "artikelrevision": "artikelrevision",
    "extra benämning": "extra_benamning",
    "extra benamning": "extra_benamning",
    "extra_benamning": "extra_benamning",
    "ritningsnummer": "ritningsnummer",
    "ritningsrevision": "ritningsrevision",
    "is_active": "is_active",
}

_COLUMN_MAP = _SWEDISH_TO_DB

_INT_FIELDS: set[str] = set()
_DECIMAL_FIELDS = {"standardpris", "saldo_varde", "nettovikt_varde"}
_BOOL_FIELDS = {"saldohanteras", "fast_vikt", "is_active"}
_SKIP_FIELDS = {
    "artikel_typ_id", "artikel_kategori_id", "artikel_kod_id",
    "varugrupp_id", "artikel_status_id", "enhet_id",
}


def _coerce_value(field: str, raw: str) -> Any:
    """Convert a raw CSV/XLSX string value to the correct Python type."""
    val = raw.strip() if isinstance(raw, str) else str(raw).strip() if raw is not None else ""
    if not val:
        return None

    if field in _SKIP_FIELDS:
        return None

    if field in _INT_FIELDS:
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return None

    if field in _DECIMAL_FIELDS:
        numeric_part = re.sub(r"[^\d,.\-]", "", val).replace(",", ".")
        if not numeric_part:
            return None
        try:
            return Decimal(numeric_part)
        except (InvalidOperation, ValueError):
            return None

    if field in _BOOL_FIELDS:
        lower = val.lower()
        if lower in ("1", "true", "yes", "ja", "sant"):
            return True
        if lower in ("0", "false", "no", "nej", "falskt"):
            return False
        return None

    return val if val else None


def _parse_csv(content: str) -> list[dict[str, str]]:
    """Parse CSV text into a list of row dicts with lowercased keys.

    FIXED (#19): delimiter is detected before DictReader is constructed.
    The original code checked reader.fieldnames immediately after construction
    which is always None at that point — so the comma fallback always fired,
    making the semicolon delimiter dead code.
    """
    # Sniff the first 2 KB to detect the delimiter before building the reader
    sample = content[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,	")
        delimiter = dialect.delimiter
    except csv.Error:
        # If sniffer cannot determine delimiter, count occurrences to decide
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","

    reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)

    rows: list[dict[str, str]] = []
    for row in reader:
        normalized = {k.strip().lower(): v for k, v in row.items() if k}
        rows.append(normalized)
    return rows


def _parse_xlsx(raw_bytes: bytes) -> list[dict[str, str]]:
    """Parse XLSX bytes into a list of row dicts with lowercased keys.

    Handles ERP System exports where the header row is not row 1.
    Scans the first 30 rows for a row containing 'Artikelnummer'.

    FIXED (#17): openpyxl is now a direct dependency in pyproject.toml.
    The ImportError guard masked a packaging error and is removed.
    """
    from openpyxl import load_workbook  # direct dep — always available

    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise AppError("Workbook has no active sheet")

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_idx = -1
    for i, row in enumerate(all_rows[:30]):
        cell_values = [str(c).strip().lower() for c in row if c]
        if "article_number" in cell_values:
            header_idx = i
            break

    if header_idx == -1:
        found_text = [
            str(c).strip()
            for row in all_rows[:10]
            for c in row
            if c and str(c).strip()
        ]
        if any(k in found_text for k in ("Kund", "Namn")):
            raise AppError(
                "Wrong file — this looks like a Customer List (customer file). "
                "Please upload it via 'Import Customers' instead."
            )
        raise AppError(
            "Could not find the 'Artikelnummer' header column in the first 30 rows. "
            "Make sure you are uploading the ERP System Article Catalogue export."
        )

    raw_headers = all_rows[header_idx]
    headers: list[str] = []
    for h in raw_headers:
        cleaned = str(h).strip().lower() if h else ""
        headers.append(cleaned)

    mapped_cols: list[tuple[int, str]] = []
    for idx, hdr in enumerate(headers):
        if hdr and hdr in _SWEDISH_TO_DB:
            mapped_cols.append((idx, _SWEDISH_TO_DB[hdr]))

    rows: list[dict[str, str]] = []
    for row_values in all_rows[header_idx + 1:]:
        if not any(row_values):
            continue
        row_dict: dict[str, str] = {}
        for col_idx, db_col in mapped_cols:
            if col_idx < len(row_values):
                val = row_values[col_idx]
                row_dict[db_col] = str(val) if val is not None else ""
        if row_dict.get("article_number", "").strip():
            rows.append(row_dict)

    return rows


class ArticleImportService:
    """Handles parsing and upserting article data from CSV/XLSX files."""

    def __init__(self, db: AsyncSession) -> None:
        self._db = db

    async def import_articles(
        self,
        *,
        csv_content: str | None = None,
        xlsx_bytes: bytes | None = None,
    ) -> ArticleImportResponse:
        """Parse and upsert articles from CSV text or XLSX bytes.

        Args:
            csv_content: UTF-8 decoded CSV text (semicolon or comma delimited).
            xlsx_bytes: Raw XLSX file bytes.

        Returns:
            ArticleImportResponse with counts and non-fatal errors.
        """
        if csv_content:
            try:
                rows = _parse_csv(csv_content)
            except Exception as exc:
                raise AppError(f"Failed to parse CSV: {exc}") from exc
        elif xlsx_bytes:
            try:
                rows = _parse_xlsx(xlsx_bytes)
            except AppError:
                raise
            except Exception as exc:
                raise AppError(f"Failed to parse XLSX: {exc}") from exc
        else:
            raise AppError("No file content provided")

        if not rows:
            return ArticleImportResponse(imported=0, skipped=0, updated=0, errors=[])

        first_row_keys = set(rows[0].keys())
        if "article_number" not in first_row_keys:
            if any(k in first_row_keys for k in ("kund", "namn")):
                raise AppError(
                    "Wrong file — this looks like a Customer List (customer file). "
                    "Please upload it via 'Import Customers' instead."
                )
            raise AppError(
                "Wrong file — could not find an 'Artikelnummer' column. "
                f"Found columns: {sorted(first_row_keys)}. "
                "Make sure you are uploading the ERP System Article Catalogue export."
            )

        imported = 0
        updated = 0
        skipped = 0
        errors: list[str] = []

        try:
            result = await self._db.execute(
                select(Article.id, Article.article_number),
            )
            existing_map: dict[str, int] = {
                row.article_number: row.id for row in result.all()
            }
        except Exception as exc:
            raise AppError(f"Failed to load existing articles: {exc}") from exc

        BATCH_SIZE = 500
        batch_count = 0

        for idx, row in enumerate(rows, start=2):
            art_nr = (row.get("article_number") or "").strip()
            if not art_nr:
                skipped += 1
                continue

            try:
                parsed: dict[str, Any] = {}
                for csv_col, model_attr in _COLUMN_MAP.items():
                    if csv_col in row and csv_col != "article_number":
                        val = _coerce_value(model_attr, row[csv_col])
                        if val is not None:
                            parsed[model_attr] = val

                if art_nr in existing_map:
                    if parsed:
                        from sqlalchemy import update
                        await self._db.execute(
                            update(Article)
                            .where(Article.id == existing_map[art_nr])
                            .values(**parsed)
                        )
                        updated += 1
                    else:
                        skipped += 1
                else:
                    # Use parsed dict — it is keyed by DB column name ("artikelbenamning")
                    # after mapping through _COLUMN_MAP. The raw row key may contain
                    # Swedish characters (e.g. "article_name" with ä) which would
                    # cause row.get("artikelbenamning") to always return None.
                    name = str(parsed.get("artikelbenamning") or "").strip()
                    if not name:
                        # Fallback: also try common Swedish variants directly from row
                        for key in ("artikelbenamning", "article_name", "article_name"):
                            name = str(row.get(key) or "").strip()
                            if name:
                                break
                    if not name:
                        errors.append(
                            f"Row {idx}: article_number '{art_nr}' skipped — "
                            f"missing artikelbenamning",
                        )
                        skipped += 1
                        continue

                    # Remove artikelbenamning from parsed to avoid
                    # "multiple values for keyword argument" error
                    parsed.pop("artikelbenamning", None)
                    article = Article(
                        article_number=art_nr,
                        artikelbenamning=name,
                        **parsed,
                    )
                    self._db.add(article)
                    existing_map[art_nr] = -1
                    imported += 1

                batch_count += 1
                if batch_count >= BATCH_SIZE:
                    await self._db.flush()
                    batch_count = 0

            except Exception as exc:
                errors.append(f"Row {idx} (article_number={art_nr}): {exc}")
                skipped += 1
                continue

        try:
            await self._db.flush()
        except Exception as exc:
            raise AppError(f"Failed to commit article import: {exc}") from exc

        logger.info(
            "article_import_complete",
            imported=imported,
            updated=updated,
            skipped=skipped,
            errors_count=len(errors),
        )

        return ArticleImportResponse(
            imported=imported,
            updated=updated,
            skipped=skipped,
            errors=errors,
        )